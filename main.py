# coding=UTF-8

import yaml
from datetime import datetime
from os import listdir, path, replace, system, remove
import openpyxl
from zipfile import BadZipfile
from pandas import read_excel
import sqlite3
import time
import traceback
from sys import exit


def load_config():
    """
    讀取 config.yaml
    """
    try:
        with open('config.yaml', 'r', encoding='utf-8') as stream:
            config = yaml.safe_load(stream)
    except FileNotFoundError:
        input("[ERROR] config.yaml not found!\n"
              "press Enter key to close this window...")
        exit()
    return config

def validate_config(config):
    """
    檢查 config.yaml 各主要 key 是否健全沒缺失
    """
    keys_should_be_exist = [
        'DIRECTORY_PATH',
        'COMBINED_SHEET_NAME',
        'SHEET_FIRST_ROW',
        'DB_NAME',
        'SQL_COMMANDS'
    ]
    for key in keys_should_be_exist:
        if key not in config.keys():
            print('[ERROR] main key missing! please check key: {key} is exist in config.yaml'.format(
                key=key))
            input("press Enter key to close this window...")
            exit()

def _past_delivery_data_to_excel(directory_path:str, combined_sheet_name:str, sheet_first_row:list):
    """
    將資料夾內過去 60 天 mail 中已經配送過的資料整理成可匯入 SQLite 的格式
    """
    print("[INFO] Begin _past_delivery_data_to_excel")
    first_row = None
    combined_sheet = openpyxl.Workbook()
    combined_sheet.create_sheet()
    if(not path.isdir(directory_path)):
        input("[ERROR] directory {} not found!\n"
              "press Enter key to close this window...".format(directory_path))
        exit()
    combined_sheet.active.append((cell for cell in sheet_first_row))
    for file_name in listdir(directory_path):
        if(file_name == ".DS_Store"):
            continue  # 跳過這種自動產生的檔案
        if("imported_" in file_name):
            continue  # 跳過已經匯入過的檔案
        try:
            read_file = openpyxl.load_workbook(path.join(directory_path, file_name))
        except BadZipfile:
            input("[ERROR] file {file_name} still open, close it first!\n"
              "press Enter key to close this window...".format(file_name=file_name))
            exit()
        sheet = read_file.active
        row_cnt = 0
        for row in sheet.rows:
            if(row[0].value == None):  # 如果該行為空則略過
                row_cnt+=1
                continue
            elif(row[0].value == "序号"):  # 如果判斷為第一列的title則要看是否已經有了就不在加入
                row_cnt+=1
                continue
            else:
                combined_sheet.active.append((cell.value for cell in row))
        # 此檔案匯入結束，於檔名前面加入 prefix
        replace(path.join(directory_path, file_name), path.join(directory_path, "imported_"+file_name))
    combined_sheet.save(combined_sheet_name + ".xlsx")
    print("[INFO] End of _past_delivery_data_to_excel")

def _consumable_levels_data_to_excel(directory_path:str, combined_sheet_name:str, sheet_first_row:list):
    """
    將資料夾內客戶閥值資料整理成可匯入 SQLite 的格式
    每個檔案需先取出第二列的時間，並新增一叫做 content_date 的 column 並全填上剛取到的時間
    然後刪除第一與第二列
    """
    print("[INFO] Begin _consumable_levels_data_to_excel")
    combined_sheet = openpyxl.Workbook()
    combined_sheet.create_sheet()
    if(not path.isdir(directory_path)):
        input("[ERROR] directory {} not found!\n"
              "press Enter key to close this window...".format(directory_path))
        exit()
    
    combined_sheet.active.append((cell for cell in sheet_first_row))
    for file_name in listdir(directory_path):
        if(file_name == ".DS_Store"):
            continue  # 跳過這種自動產生的檔案
        if("imported_" in file_name):
            continue  # 跳過已經匯入過的檔案
        try:
            read_file = openpyxl.load_workbook(path.join(directory_path, file_name))
        except BadZipfile:
            input("[ERROR] file {file_name} still open, close it first!\n"
              "press Enter key to close this window...".format(file_name=file_name))
            exit()
        sheet = read_file.active
        row_cnt = 0
        report_content_date = 0
        for row in sheet.rows:
            if(row[0].value == None or row_cnt == 0 or row_cnt == 2):  
                # 如果該行為空略過
                # 第一列是產 report 的時間，不需要故略過
                # 第三列是 column header，不需要故略過
                row_cnt+=1
            elif(row_cnt == 1):  # 第二列需要取時間
                row_cnt+=1
                report_content_date = row[1].value
            else:
                row_cnt+=1
                combined_sheet.active.append((report_content_date,) + (tuple)(cell.value for cell in row))
        # 此檔案匯入結束，於檔名前面加入 prefix
        replace(path.join(directory_path, file_name), path.join(directory_path, "imported_"+file_name))
    combined_sheet.save(combined_sheet_name + ".xlsx")
    print("[INFO] End of _consumable_levels_data_to_excel")

def get_need_refill_sheet(command):
    """
    取 view table
    """
    print("[INFO] Begin get_need_refill_sheet")
    csv_file_name = datetime.now().strftime("%Y-%m-%d") + "_need_refill.csv"
    if(path.isfile(csv_file_name)):
        remove(csv_file_name)
        print("[INFO] 舊有表格" + csv_file_name + " 表格已刪除")
    system(command)
    print("[INFO] " + csv_file_name + " 表格已生成")
    print("[INFO] End of get_need_refill_sheet")

def apply_settings(db_conn):
    """
    重取 config 中的 threshold 數值
    """
    print("[INFO] Begin apply settings")
    config = load_config()
    validate_config(config)
    global LEVEL_THRESHOLD
    LEVEL_THRESHOLD = config.get('LEVEL_THRESHOLD')
    print("[INFO] End of apply settings")
    
def import_past_deliver(db_conn):
    print("[INFO] Begin import past devier data...")
    # 將資料夾內過去 60 天 mail 中已經配送過的資料整理成可匯入 SQLite 的格式
    _past_delivery_data_to_excel(PAST_DELIVERY_DATA_PATH, COMBINED_PAST_DELIVERY_SHEET_NAME, PAST_DELIVERY_SHEET_FIRST_ROW)
    excel_file = read_excel(COMBINED_PAST_DELIVERY_SHEET_NAME + ".xlsx",
                                        sheet_name='Sheet',
                                        header=0)
    #將通整好的 excel 匯入至 SQLite 中
    excel_file.to_sql('past_delivery_data', db_conn, if_exists='append', index=False)
    # 用已知的 goods_type 去填補 past_delivery_data 中 goods_type 為 NULL 的欄位
    cur = db_conn.cursor()
    fill_known_goods_type = SQL_COMMANDS.get('fill_known_goods_type')
    cur.execute(fill_known_goods_type)
    db_conn.commit()
    cur.close()
    if(path.isfile(COMBINED_PAST_DELIVERY_SHEET_NAME + ".xlsx")):
        remove(COMBINED_PAST_DELIVERY_SHEET_NAME + ".xlsx")
    print("[INFO] End of import past devier data...")

def import_consumable_levels(db_conn):
    print("[INFO] Begin import consumable levels data...")
    # 將資料夾內客戶閥值資料整理成可匯入 SQLite 的格式
    _consumable_levels_data_to_excel(CONSUMABLE_LEVELS_DATA_PATH, COMBINED_CONSUMABLE_LEVELS_SHEET_NAME, CONSUMABLE_LEVELS_SHEET_FIRST_ROW)
    excel_file = read_excel(COMBINED_CONSUMABLE_LEVELS_SHEET_NAME + ".xlsx",
                                        sheet_name='Sheet',
                                        header=0)
    #將通整好的 excel 匯入至 SQLite 中
    excel_file.to_sql('consumable_levels', db_conn, if_exists='append', index=False)
    if(path.isfile(COMBINED_CONSUMABLE_LEVELS_SHEET_NAME + ".xlsx")):
        remove(COMBINED_CONSUMABLE_LEVELS_SHEET_NAME + ".xlsx")
    print("[INFO] End of  import consumable levels data...")

def insert_or_ignore_deliver_status(db_conn):
    print("[INFO] Begin import insert_or_ignore_deliver_status...")
    cur = db_conn.cursor()
    for info in COLUMN_NAME_AND_GOODS_TYPE:
        column_name = info[0]
        goods_type = info[1]
        insert_or_ignore_deliver_status = SQL_COMMANDS.get('insert_or_ignore_deliver_status_with_2args').format(
            goods_type=goods_type, 
            column_name=column_name
        )
        cur.execute(insert_or_ignore_deliver_status)
    db_conn.commit()
    cur.close()
    print("[INFO] End of import insert_or_ignore_deliver_status...")

def update_product_level(db_conn):
    """
    更新產品用量
    """
    print("[INFO] Begin update_product_level...")
    cur = db_conn.cursor()
    for info in COLUMN_NAME_AND_GOODS_TYPE:
        column_name = info[0]
        goods_type = info[1]
        update_product_level = SQL_COMMANDS.get('update_product_level_with_2args').format(
            goods_type=goods_type, 
            column_name=column_name
        )
        cur.execute(update_product_level)
    db_conn.commit()
    cur.close()
    print("[INFO] End of update_product_level...")

def first_time_update_deliver_status(db_conn):
    """
    第一次執行，需排除掉 past_delivery 中的機器不派送
    更新派送狀態並確認如果已經填充過就 reset need_refill_count 避免重複派送
    """
    print("[INFO] Begin first_time_update_deliver_status...")
    cur = db_conn.cursor()
    get_all_serial_number = SQL_COMMANDS.get('get_all_serial_number')
    serial_numbers = cur.execute(get_all_serial_number).fetchall()
    for info in COLUMN_NAME_AND_GOODS_TYPE:
        for serial_number in serial_numbers:
            column_name = info[0]
            goods_type = info[1]
            first_time_update_deliver_status = SQL_COMMANDS.get('first_time_update_deliver_status_with_3args').format(
            column_name=column_name,
            serial_number=serial_number[0], 
            goods_type=goods_type)
            cur.execute(first_time_update_deliver_status)
            reset_need_refill_count =  SQL_COMMANDS.get('reset_need_refill_count_with_3args').format(
                column_name=column_name,
                serial_number=serial_number[0],
                goods_type=goods_type)
            cur.execute(reset_need_refill_count)
    db_conn.commit()
    cur.close()
    print("[INFO] End of first_time_update_deliver_status...")

def update_threshold(db_conn):
    """
    更新閥值
    若發現 threshold 有變動先把今天的 deliver_status 的時間洗掉
    然後用新的 threshold 跑 update_deliver_status_after_change_threshold
    """
    print("[INFO] Begin update_threshold...")
    cur = db_conn.cursor()
    update_threshold = SQL_COMMANDS.get('update_threshold_with_arg').format(threshold=LEVEL_THRESHOLD)
    cur.execute(update_threshold)
    update_result = False
    if(cur.rowcount > 0):
        print("[INFO] Begin reset_date_of_deliver_status...")
        reset_date_of_deliver_status = SQL_COMMANDS.get('reset_date_of_deliver_status')
        cur.execute(reset_date_of_deliver_status)
        update_result = True
        print("[INFO] End of reset_date_of_deliver_status...")
    cur.close()
    db_conn.commit()
    print("[INFO] End of update_threshold...")
    return update_result

def update_deliver_status_after_change_threshold(db_conn):
    """
    更換閥值後的更新派送狀態
    如果 need_refill or need_refill_count 已經大於 0 代表已經曾經寄送且還沒客戶還沒 refill
    那就算 threshold 再降低造成要寄送也不用再送了
    但如果 threshold 提高，導致有一些反而高於閥值那已經寄送過的也來不及了所以不動作XD
    如果都是 0 表示還沒寄送過，那就用新的閥值再來判斷一次看有沒有需要寄送
    """
    print("[INFO] Begin update_deliver_status_after_change_threshold...")
    cur = db_conn.cursor()
    get_all_serial_number = SQL_COMMANDS.get('get_all_serial_number')
    serial_numbers = cur.execute(get_all_serial_number).fetchall()
    for info in COLUMN_NAME_AND_GOODS_TYPE:
        for serial_number in serial_numbers:
            column_name = info[0]
            goods_type = info[1]
            change_threshold_update_deliver_status = SQL_COMMANDS.get('change_threshold_update_deliver_status_with_3args').format(
            column_name=column_name,
            serial_number=serial_number[0], 
            goods_type=goods_type)
            cur.execute(change_threshold_update_deliver_status)
            reset_need_refill_count =  SQL_COMMANDS.get('reset_need_refill_count_with_3args').format(
                column_name=column_name,
                serial_number=serial_number[0],
                goods_type=goods_type)
            cur.execute(reset_need_refill_count)
    db_conn.commit()
    cur.close()
    print("[INFO] End of update_deliver_status_after_change_threshold...")

def update_deliver_status(db_conn):
    """
    更新派送狀態並確認如果已經填充過就 reset need_refill_count 避免重複派送
    """
    print("[INFO] Begin update_deliver_status...")
    cur = db_conn.cursor()
    get_all_serial_number = SQL_COMMANDS.get('get_all_serial_number')
    serial_numbers = cur.execute(get_all_serial_number).fetchall()
    for info in COLUMN_NAME_AND_GOODS_TYPE:
        for serial_number in serial_numbers:
            column_name = info[0]
            goods_type = info[1]
            update_deliver_status = SQL_COMMANDS.get('update_deliver_status_with_3args').format(
            column_name=column_name,
            serial_number=serial_number[0], 
            goods_type=goods_type)
            cur.execute(update_deliver_status)
            reset_need_refill_count =  SQL_COMMANDS.get('reset_need_refill_count_with_3args').format(
                column_name=column_name,
                serial_number=serial_number[0],
                goods_type=goods_type)
            cur.execute(reset_need_refill_count)
    db_conn.commit()
    cur.close()
    print("[INFO] End of update_deliver_status...")

def get_deliver_status(db_conn):
    print("[INFO] begin update and get deliver status")
    # 更新 past_delivery_data
    import_past_deliver(db_conn)
    # 更新 consumable_level
    import_consumable_levels(db_conn)
    # 新增舊的 deliver_status table 中沒有的機器
    insert_or_ignore_deliver_status(db_conn)
    if update_threshold(db_conn):
        # 有變更閥值故執行有變更閥值的 update_deliver_status 邏輯
        update_deliver_status_after_change_threshold(db_conn)
    else:
        # 趁此時 consumable_levels 的值是今天的，而 product_level 的值是昨天的
        # 來設定 deliver_status need_refill 的值
        update_deliver_status(db_conn)
        # 將今天的數值填進 product_level 中
        update_product_level(db_conn)
    # 取 view table 結束這回合
    get_need_refill_sheet(export_query_command)
    print("[INFO] End of update and deliver status")

def first_time_get_deliver_status(db_conn):
    print("[INFO] begin update and get deliver status")
    # 更新 past_delivery_data
    import_past_deliver(db_conn)
    # 更新 consumable_level
    import_consumable_levels(db_conn)
    # 新增舊的 deliver_status table 中沒有的機器
    insert_or_ignore_deliver_status(db_conn)
    # 將閥值寫入
    update_threshold(db_conn)
    # 將今天的數值填進 product_level 中
    update_product_level(db_conn)
    # 來設定 deliver_status need_refill 的值
    first_time_update_deliver_status(db_conn)
    # 取 view table 結束這回合
    get_need_refill_sheet(export_query_command)
    print("[INFO] End of update and deliver status")

def reset_db(db_conn):
    print("[INFO] Begin db reset...")
    db_conn.close()
    remove(DB_NAME)
    f = open(DB_NAME, "a")
    f.close()
    new_db_conn = sqlite3.connect(DB_NAME)
    cur = new_db_conn.cursor()
    create_db_commands = SQL_COMMANDS.get('reset_db_commands')
    for key in create_db_commands.keys():
        sql_command = create_db_commands.get(key)
        print("[INFO] proccessing {command}".format(command=key))
        cur.execute(sql_command)
    cur.close()
    new_db_conn.commit()
    print("[INFO] End of db reset...")

def exit_program(arg):
    input("exit program, press Enter to close window....")
    exit()

def test(db_conn):
    update_deliver_status(db_conn)

if __name__ == '__main__':
    config = load_config()
    validate_config(config)
    LEVEL_THRESHOLD = config.get('LEVEL_THRESHOLD')
    PAST_DELIVERY_DATA_PATH = config.get('DIRECTORY_PATH').get('past_delivery_data')
    CONSUMABLE_LEVELS_DATA_PATH = config.get('DIRECTORY_PATH').get('consumable_levels_data')
    COMBINED_PAST_DELIVERY_SHEET_NAME = config.get('COMBINED_SHEET_NAME').get('past_delivery')
    COMBINED_CONSUMABLE_LEVELS_SHEET_NAME = config.get('COMBINED_SHEET_NAME').get('consumable_levels')
    PAST_DELIVERY_SHEET_FIRST_ROW = config.get('SHEET_FIRST_ROW').get('past_delivery_data')
    CONSUMABLE_LEVELS_SHEET_FIRST_ROW = config.get('SHEET_FIRST_ROW').get('consumable_levels_data')
    DB_NAME = config.get('DB_NAME')
    SQL_COMMANDS = config.get('SQL_COMMANDS')
    export_query_command = config.get('export_query_command_with_arg').format(date=datetime.now().strftime("%Y-%m-%d"))
    COLUMN_NAME_AND_GOODS_TYPE = [
        ("Black Level", "黑"),
        ("Cyan Level", "藍"),
        ("Magenta Level", "紅"),
        ("Yellow Level", "黃"),
        ("Black Drum Level", "黑鼓"),
        ("Cyan Drum Level", "藍鼓"),
        ("Magenta Drum Level", "紅鼓"),
        ("Yellow Drum Level", "黃鼓"),
        ("Drum Kit Level", "drum kit"),
        ("Transfer Kit Level", "transfer kit"),
        ("Fuser Kit Level", "fuser kit"),
        ("Cleaning Kit Level", "clean kit"),
        ("Maintenance Combo Kit Level", "maintainace combo kit"),
        ("Document Feeder Kit Level", "document feeder kit"),
        ("Roller Level", "roller")
    ]
    function_list = {
        "0": first_time_get_deliver_status,
        "1": get_deliver_status,
        "2": apply_settings,
        "3": reset_db,  # delete all tables and views
        "4": exit_program
    }

    while True:
        choose = input("請輸入功能代號:\n"
        "0. 第一次執行更新並取得應派送設備表單(將會排除曾經派送清單中的機器)\n"
        "1. 更新並取得應派送設備表單\n"
        "(將會匯入檔案且過檔名將會改成 imported_ 開頭，並無法再次匯入)\n"
        "2. 套用新設定(閥值)，需要再使用功能 1 才能產出根據新閥值計算的設備表單\n"
        "3. 刪除AND創建全新DB(!注意!原 DB 將被完整刪除，資料將不能恢復)\n"
        "4. 退出程式\n"
        "選擇: ")
        db_conn = sqlite3.connect(DB_NAME)
        func = function_list.get(choose, lambda x: print("Invalid number!\n\n"))
        if func is reset_db:
            while True:
                delete = input("此選擇將會刪除整個 DB 並另起一個新的資料庫\n"
                "過去的紀錄將全部消失，此操作將無法還原，請問要繼續嗎?\ny/n\n"
                "選擇: ")
                if delete == "n" or delete == "N":
                    exit()
                if delete == "y" or delete == "Y":
                    break
        try:
            func(db_conn)  # 執行選取的邏輯
            db_conn.close()
        except  Exception as e:
            traceback.print_exc()
            input("exception occur!\n"
              "press Enter key to close this window...")
            exit()
        
